import imaplib
import base64
import os
import time
from configparser import ConfigParser, ExtendedInterpolation
import wget
import urllib
import xlrd
import uuid
import csv
from bson.objectid import ObjectId
import json
from datetime import datetime
import requests
from difflib import get_close_matches
from requests import post, get, delete
import sys
import time
import xlwt
import xlutils
from xlutils.copy import copy
import shutil
import re
from xlrd import open_workbook
from xlutils.copy import copy as xl_copy
import logging
import logging.handlers
import time
from logging.handlers import TimedRotatingFileHandler
import xlsxwriter
import argparse
import sys
from os import path
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import gdown

# get current working directory
currentDirectory = os.getcwd()

# Read config file 
config = ConfigParser()
config.read('common_config/config.ini')

# email regex
regex = "\"?([-a-zA-Z0-9.`?{}]+@\w+\.\w+)\"?"

criteriaLookUp = dict()
millisecond = None
programNameInp = None
environment = None
observationId = None
solutionName = None
pointBasedValue = None
entityType = None
allow_multiple_submissions = None
scopeEntityType = ""
programName = None
userEntity = None
roles = ""
dictCritLookUp = {}
isProgramnamePresent = None
solutionLanguage = None
keyWords = None
entityTypeId = None
solutionDescription = None
creator = None
dikshaLoginId = None
criteriaName = None
solutionId = None
API_log = None
listOfFoundRoles = []
entityToUpload = None
programID = None
programExternalId = None
programDescription = None
criteriaLookUp = dict()
themesSheetList = []
themeRubricFileObj = dict()
criteriaLevelsReport = False
ecm_sections = dict()
criteriaLevelsCount = 0
numberOfResponses = 0
criteriaIdNameDict = dict()
criteriaLevels = list()
matchedShikshalokamLoginId = None
scopeEntities = []
scopeRoles = []
countImps = 0
ecmToSection = dict()
entitiesPGMID = []
OrgName = []
solutionRolesArr = []
startDateOfResource = None
endDateOfResource = None
solutionRolesArray = []
solutionStartDate = ""
solutionEndDate = ""
projectCreator = ""
orgIds = []
ccRootOrgName = None
ccRootOrgId  = None
certificatetemplateid = None



def terminatingMessage(msg):
    print(msg)
    sys.exit()




def validateSheets(filePathAddObs, accessToken, parentFolder):
    wbObservation1 = xlrd.open_workbook(filePathAddObs, on_demand=True)
    sheetNames1 = wbObservation1.sheet_names()
    survey_sheet_names = ['Instructions', 'details', 'questions']

    if (len(survey_sheet_names) == len(sheetNames1)) and ((set(survey_sheet_names) == set(sheetNames1))):
        print("--->Survey file detected.<---")
    else:
        terminatingMessage("Please check the Input sheet.")

def apicheckslog(solutionName_for_folder_path, messageArr):
    file_exists = solutionName_for_folder_path + '/apiHitLogs/apiLogs.csv'
    # global fileheader
    fileheader = ["Resource","Process","Status","Remark"]
    # with open(file_exists, 'w',newline='') as file:
    #      writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',')
    #      writer.writerows([fileheader])

    if not path.exists(file_exists):
        with open(file_exists, 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',')
            writer.writerows([fileheader])

    with open(file_exists, 'a', newline='') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',')
        writer.writerows([messageArr])

def courseMapToProgram(accessToken, courseLink, parentFolder):
    terminatingMessage("---> Course not part of the product ...")


# program creation function 
def programCreation(accessToken, parentFolder, externalId, pName, pDescription, keywords, entities, roles, orgIds,creatorKeyCloakId, creatorName,entitiesPGM,mainRole,rolesPGM):
    messageArr = []
    messageArr.append("++++++++++++ Program Creation ++++++++++++")
    # program creation url 
    programCreationurl = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'programCreationurl')
    messageArr.append("Program Creation URL : " + programCreationurl)
    # program creation payload
    payload = json.dumps({
        "externalId": externalId,
        "name": pName,
        "description": pDescription,
        "resourceType": [
            "program"
        ],
        "language": [
            "English"
        ],
        "keywords": keywords,
        "concepts": [],
        "createdFor": orgIds,
        "rootOrganisations": orgIds,
        "startDate": startDateOfProgram,
        "endDate": endDateOfProgram,
        "imageCompression": {
            "quality": 10
        },
        "creator": creatorName,
        "owner": creatorKeyCloakId,
        "author": creatorKeyCloakId,
        "scope": {
            "entityType": scopeEntityType,
            "entities": entitiesPGMID,
            "roles": roles
        },
        "metaInformation": {
            "state":entitiesPGM.split(","),
            "roles": mainRole.split(",")
            },
            "requestForPIIConsent":True
            })

    messageArr.append("Body : " + str(payload))
    headers = {'X-authenticated-user-token': accessToken,
               'internal-access-token': config.get(environment, 'internal-access-token'),
               'Content-Type': 'application/json',
               'Authorization':config.get(environment, 'Authorization')}
    
    # program creation 
    responsePgmCreate = requests.request("POST", programCreationurl, headers=headers, data=(payload))
    messageArr.append("Program Creation Status Code : " + str(responsePgmCreate.status_code))
    messageArr.append("Program Creation Response : " + str(responsePgmCreate.text))
    messageArr.append("Program body : " + str(payload))

    # save logs 
    createAPILog(parentFolder, messageArr)
    # check status 
    fileheader = [pName, ('Program Sheet Validation'), ('Passed')]
    createAPILog(parentFolder, messageArr)
    apicheckslog(parentFolder, fileheader)
    if responsePgmCreate.status_code == 200:
        responsePgmCreateResp = responsePgmCreate.json()
    else:
        # terminate execution
        terminatingMessage("Program creation API failed. Please check logs.")


# this function is used to create the sheet of PDPM for API requerment
def programmappingpdpmsheetcreation(MainFilePath,accessToken, program_file,programexternalId,parentFolder):
    pdpmsheet = MainFilePath+ "/pdpmmapping/"
    if not os.path.exists(pdpmsheet):
        os.mkdir(pdpmsheet)
    print(program_file)
    print(MainFilePath)

    wbproject = xlrd.open_workbook(program_file, on_demand=True)
    projectSheetNames = wbproject.sheet_names()

    mappingsheet = wbproject.sheet_by_name('Program Details')
    keysProject = [mappingsheet.cell(1, col_index_env).value for col_index_env in
                   range(mappingsheet.ncols)]

    pdpmcolo1 = ["user","role","entity","entityOperation","keycloak-userId","acl_school","acl_cluster","programOperation",
                "platform_role","programs","_arrayFields"]
    with open(pdpmsheet + 'mapping.csv', 'w') as file:
         writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
         writer.writerows([pdpmcolo1])

    wbPgm = xlrd.open_workbook(program_file, on_demand=True)
    global programNameInp
    sheetNames = wbPgm.sheet_names()
    for sheetEnv in sheetNames:
        if sheetEnv == "Instructions":
            pass
        elif sheetEnv.strip().lower() == 'program details':
            print("--->Checking Program details sheet...")
            detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
            keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                       range(detailsEnvSheet.ncols)]
            for row_index_env in range(2, detailsEnvSheet.nrows):
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                  for
                                  col_index_env in range(detailsEnvSheet.ncols)}
                programNameInp = dictDetailsEnv['Title of the Program'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Title of the Program'] else terminatingMessage("\"Title of the Program\" must not be Empty in \"Program details\" sheet")

            extIdPGM = dictDetailsEnv['Program ID'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Program ID'] else terminatingMessage("\"Program ID\" must not be Empty in \"Program details\" sheet")

            programdesigner = dictDetailsEnv['Diksha username/user id/email id/phone no. of Program Designer'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Program ID'] else terminatingMessage("\"Diksha username/user id/email id/phone no. of Program Designer\" must not be Empty in \"Program details\" sheet")
            userDetails = fetchUserDetails(environment, accessToken, programdesigner)
            
            creatorKeyCloakId = userDetails[0]
            creatorName = userDetails[1]
            if "PROGRAM_DESIGNER" in userDetails[3]:
                creatorKeyCloakId = userDetails[0]
                creatorName = userDetails[1]
            else :
                terminatingMessage("user does't have program designer role")

            pdpmcolo1 = [creatorName, " ", " ", " ", creatorKeyCloakId, " ", " ","ADD","PROGRAM_DESIGNER", extIdPGM, "programs"]
            with open(pdpmsheet + 'mapping.csv', 'a') as file:
                writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                writer.writerows([pdpmcolo1])
                fileheader = [creatorName,"program designer mapped successfully","Passed"]
                apicheckslog(parentFolder,fileheader)


        elif sheetEnv.strip().lower() == 'program manager details':
            print("--->Program Manager Details...")
            detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
            keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                       range(detailsEnvSheet.ncols)]
            for row_index_env in range(2, detailsEnvSheet.nrows):
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                  for
                                  col_index_env in range(detailsEnvSheet.ncols)}

                if str(dictDetailsEnv['Is a SSO user?']).strip() == "YES":
                    programmanagername2 = dictDetailsEnv['Diksha user id ( profile ID)'] if dictDetailsEnv['Diksha user id ( profile ID)'] else terminatingMessage("\"Diksha user id ( profile ID)\" must not be Empty in \"Program details\" sheet")
                else:
                    try :
                        programmanagername2 = dictDetailsEnv['Login ID on DIKSHA'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Login ID on DIKSHA'] else terminatingMessage("\"Login ID on DIKSHA\" must not be Empty in \"Program details\" sheet")
                        userDetails = fetchUserDetails(environment, accessToken, programmanagername2)
                    except :
                        programmanagername2 = dictDetailsEnv['Diksha user id ( profile ID)'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Diksha user id ( profile ID)'] else terminatingMessage("\"Diksha user id ( profile ID)\" must not be Empty in \"Program details\" sheet")
                        userDetails = fetchUserDetails(environment, accessToken, programmanagername2)

                userDetails = fetchUserDetails(environment, accessToken, programmanagername2)
                creatorKeyCloakId = userDetails[0]
                creatorName = userDetails[1]
                if "PROGRAM_MANAGER" in userDetails[3]:
                    creatorKeyCloakId = userDetails[0]
                    creatorName = userDetails[1]
                else:
                    terminatingMessage("user does't have program manager role")

                pdpmcolo1 = [creatorName, " ", " ", " ", creatorKeyCloakId, " ", " ","ADD","PROGRAM_MANAGER", extIdPGM, "programs"]

                with open(pdpmsheet + 'mapping.csv', 'a') as file:
                    writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',lineterminator='\n')
                    writer.writerows([pdpmcolo1])
                messageArr.append("Response : " + str(pdpmcolo1))
                createAPILog(parentFolder, messageArr)

                fileheader = [creatorName,"program manager mapped succesfully","Passed"]
                apicheckslog(parentFolder,fileheader)



# this function is used for call the api and map the pdpm roles which we created
def Programmappingapicall(MainFilePath,accessToken, program_file,parentFolder):
    urlpdpmapi = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'Pdpmurl')
    headerpdpmApi = {
        'Authorization': config.get(environment, 'Authorization'),
        'X-authenticated-user-token': accessToken,
        'X-Channel-id': config.get(environment, 'X-Channel-id'),
        'internal-access-token': config.get(environment, 'internal-access-token')
    }
    payload = {}
    filesProject = {
        'userRoles': open(MainFilePath + '/pdpmmapping/mapping.csv', 'rb')
    }

    responseProgrammappingApi = requests.post(url=urlpdpmapi, headers=headerpdpmApi,
                                             data=payload,
                                             files=filesProject)
    messageArr = ["program mapping sheet.",
                  "File path : " + MainFilePath + '/pdpmmapping/mapping.csv']
    messageArr.append("Upload status code : " + str(responseProgrammappingApi.status_code))
    createAPILog(parentFolder, messageArr)

    if responseProgrammappingApi.status_code == 200:
        print('--->program manager and designer mapping is Success')
        with open(MainFilePath + '/pdpmmapping/mappinginternal.csv', 'w+') as projectRes:
            projectRes.write(responseProgrammappingApi.text)
            messageArr.append("Response : " + str(responseProgrammappingApi.text))
            createAPILog(parentFolder, messageArr)
    else:
        messageArr.append("Response : " + str(responseProgrammappingApi.text))
        createAPILog(parentFolder, messageArr)
        fileheader = ["PDPM mapping","PDPM mapping is failed","Failed","check PDPM sheet"]
        apicheckslog(parentFolder,fileheader)
        sys.exit()


def fetchEntityId(solutionName_for_folder_path, accessToken, entitiesNameList, scopeEntityType):
    urlFetchEntityListApi = config.get(environment, 'host')+config.get(environment, 'searchForLocation')
    headerFetchEntityListApi = {
        'Content-Type': config.get(environment, 'Content-Type'),
        'Authorization': config.get(environment, 'AuthorizationForHost'),
        'X-authenticated-user-token': accessToken,
        'X-Channel-id': config.get(environment, 'X-Channel-id'),
    }
    payload = {
        "request": {
            "filters": {
                "type": scopeEntityType
            },
            "limit": 1000
        }
    }
    responseFetchEntityListApi = requests.post(url=urlFetchEntityListApi, headers=headerFetchEntityListApi,data=json.dumps(payload))
    messageArr = ["Entities List Fetch API executed.", "URL  : " + str(urlFetchEntityListApi),
                  "Status : " + str(responseFetchEntityListApi.status_code)]
    createAPILog(solutionName_for_folder_path, messageArr)
    if responseFetchEntityListApi.status_code == 200:
        responseFetchEntityListApi = responseFetchEntityListApi.json()
        entitiesLookup = dict()
        entityToUpload = list()
        for listEntities in responseFetchEntityListApi['result']['response']:
            entitiesLookup[listEntities['name'].lower().lstrip().rstrip()] = listEntities['id'].lstrip().rstrip()
        entitiesFlag = False
        for eachUserEntity in entitiesNameList:
            try:
                entityId = entitiesLookup[eachUserEntity.lower().lstrip().rstrip()]
                entitiesFlag = True
            except:
                entitiesFlag = False
            if entitiesFlag:
                entityToUpload.append(entityId)
            else:
                print("Entity Not found in DB...")
                print("Entity name : " + str(eachUserEntity))
                messageArr = ["Entity Not found : ", "URL  : " + str(eachUserEntity)]
                createAPILog(solutionName_for_folder_path, messageArr)

        messageArr = ["Entities to upload : " + str(entityToUpload)]
        createAPILog(solutionName_for_folder_path, messageArr)
        if len(entityToUpload) == 0:
            terminatingMessage("--->Scope Entity error.")
        return entityToUpload
    else:
        messageArr = ["Error in Location search",str(responseFetchEntityListApi.status_code)]
        createAPILog(solutionName_for_folder_path, messageArr)
        terminatingMessage("---> Error in location search.")

def getProgramInfo(accessTokenUser, solutionName_for_folder_path, programNameInp):
    global programID, programExternalId, programDescription, isProgramnamePresent, programName
    programName = programNameInp
    programUrl = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'fetchProgramInfoApiUrl') + programNameInp.lstrip().rstrip()
    terminatingMessage
    headersProgramSearch = {'Authorization': config.get(environment, 'Authorization'),
                            'Content-Type': 'application/json', 'X-authenticated-user-token': accessTokenUser,
                            'internal-access-token': config.get(environment, 'internal-access-token')}
    responseProgramSearch = requests.post(url=programUrl, headers=headersProgramSearch)
    messageArr = []

    messageArr.append("Program Search API")
    messageArr.append("URL : " + programUrl)
    messageArr.append("Status Code : " + str(responseProgramSearch.status_code))
    messageArr.append("Response : " + str(responseProgramSearch.text))
    createAPILog(solutionName_for_folder_path, messageArr)
    messageArr = []
    if responseProgramSearch.status_code == 200:
        print('--->Program fetch API Success')
        messageArr.append("--->Program fetch API Success")
        responseProgramSearch = responseProgramSearch.json()
        countOfPrograms = len(responseProgramSearch['result']['data'])
        messageArr.append("--->Program Count : " + str(countOfPrograms))
        if countOfPrograms == 0:
            messageArr.append("No program found with the name : " + str(programName.lstrip().rstrip()))
            messageArr.append("******************** Preparing for program Upload **********************")
            print("No program found with the name : " + str(programName.lstrip().rstrip()))
            print("******************** Preparing for program Upload **********************")
            createAPILog(solutionName_for_folder_path, messageArr)
            fileheader = ["Program name fetch","Successfully fetched program name","Passed"]
            apicheckslog(solutionName_for_folder_path,fileheader)
            return False
        else:
            getProgramDetails = []
            for eachPgm in responseProgramSearch['result']['data']:
                if eachPgm['isAPrivateProgram'] == False:
                    programID = eachPgm['_id']
                    programExternalId = eachPgm['externalId']
                    programDescription = eachPgm['description']
                    isAPrivateProgram = eachPgm['isAPrivateProgram']
                    getProgramDetails.append([programID, programExternalId, programDescription, isAPrivateProgram])
                    if len(getProgramDetails) == 0:
                        print("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                        messageArr.append("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                        createAPILog(solutionName_for_folder_path, messageArr)
                        fileheader = ["program find api is running","found"+str(len(
                            getProgramDetails))+"programs in backend","Failed","found"+str(len(
                            getProgramDetails))+"programs ,check logs"]
                        apicheckslog(solutionName_for_folder_path,fileheader)
                        terminatingMessage("Aborting...")
                    elif len(getProgramDetails) > 1:
                        print("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                        messageArr.append("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                        createAPILog(solutionName_for_folder_path, messageArr)
                        terminatingMessage("Aborting...")

                    else:
                        programID = getProgramDetails[0][0]
                        programExternalId = getProgramDetails[0][1]
                        programDescription = getProgramDetails[0][2]
                        isAPrivateProgram = getProgramDetails[0][3]
                        isProgramnamePresent = True
                        messageArr.append("programID : " + str(programID))
                        messageArr.append("programExternalId : " + str(programExternalId))
                        messageArr.append("programDescription : " + str(programDescription))
                        messageArr.append("isAPrivateProgram : " + str(isAPrivateProgram))
                    createAPILog(solutionName_for_folder_path, messageArr)
    else:
        print("Program search API failed...")
        print(responseProgramSearch)
        messageArr.append("Program search API failed...")
        createAPILog(solutionName_for_folder_path, messageArr)
        terminatingMessage("Response Code : " + str(responseProgramSearch.status_code))
    return True


# function to create API hit logs 
def createAPILog(solutionName_for_folder_path, messageArr):
    file_exists = solutionName_for_folder_path + '/apiHitLogs/apiLogs.txt'
    # check if the file existis or not and create a file 
    if not path.exists(file_exists):
        API_log = open(file_exists, "w", encoding='utf-8')
        API_log.write("===============================================================================")
        API_log.write("\n")
        API_log.write("ENVIRONMENT : " + str(environment))
        API_log.write("\n")
        API_log.write("===============================================================================")
        API_log.write("\n")
        API_log.close()

    API_log = open(file_exists, "a", encoding='utf-8')
    API_log.write("\n")
    for msg in messageArr:
        API_log.write(msg)
        API_log.write("\n")
    API_log.close()

def fetchScopeRole(solutionName_for_folder_path, accessToken, roleNameList):
    urlFetchRolesListApi = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'listOfRolesApi')
    headerFetchRolesListApi = {
        'Content-Type': config.get(environment, 'Content-Type'),
        'Authorization': config.get(environment, 'Authorization'),
        'X-authenticated-user-token': accessToken,
        'X-Channel-id': config.get(environment, 'X-Channel-id'),
    }
    responseFetchRolesListApi = requests.post(url=urlFetchRolesListApi, headers=headerFetchRolesListApi)
    rolesLookup = dict()
    rolesReturn = list()
    messageArr = ["Roles list fetch API called.", "URL  : " + str(urlFetchRolesListApi),
                  "Status Code : " + str(responseFetchRolesListApi.status_code)]
    createAPILog(solutionName_for_folder_path, messageArr)
    if responseFetchRolesListApi.status_code == 200:
        responseFetchRolesListApi = responseFetchRolesListApi.json()
        for listRoles in responseFetchRolesListApi['result']:
            eachDict = dict()
            eachDict['id'] = listRoles['_id'].lstrip().rstrip()
            eachDict['code'] = listRoles['code'].lstrip().rstrip()
            rolesLookup[listRoles['code']] = eachDict['id']
            rolesReturn.append(listRoles['code'].lstrip().rstrip())
    else:
        terminatingMessage("---> error in subroles API.")

    userRolesFromInp = roleNameList
    listOfFoundRoles = list()
    if len(userRolesFromInp) == 0:
        terminatingMessage("Roles fields must not be empty.")
    for ur in userRolesFromInp:
        rolesFlag = True
        try:
            roleDetails = rolesLookup[ur.lstrip().rstrip()]
            rolesFlag = True
        except:
            rolesFlag = False

        if rolesFlag:
            print("Role Found... : " + ur)
            listOfFoundRoles.append(ur)
        else:
            if "all" in userRolesFromInp:
                listOfFoundRoles = ["all"]
            else:
                print("Role error...")
                print("Role : " + ur)
                messageArr = ["Roles Error", "URL  : ", "Role : " + ur]
                createAPILog(solutionName_for_folder_path, messageArr)

    messageArr = ["Accepted Roles : " + str(listOfFoundRoles)]
    createAPILog(solutionName_for_folder_path, messageArr)
    if len(listOfFoundRoles) == 0:
        messageArr = ["No roles matched our DB "]
        createAPILog(solutionName_for_folder_path, messageArr)
        print("No Roles matched our DB.")
    return listOfFoundRoles

# Open and validate program sheet 
def programsFileCheck(filePathAddPgm, accessToken, parentFolder, MainFilePath):
    program_file = filePathAddPgm
    # open excel file 
    wbPgm = xlrd.open_workbook(filePathAddPgm, on_demand=True)
    global programNameInp
    sheetNames = wbPgm.sheet_names()
    # list of sheets in the program sheet 
    pgmSheets = ["Instructions", "Program Details", "Resource Details","Program Manager Details"]

    # checking the sheets in the program sheet 
    if (len(sheetNames) == len(pgmSheets)) and ((set(sheetNames) == set(pgmSheets))):
        print("--->Program Template detected.<---")
        # iterate through the sheets 
        for sheetEnv in sheetNames:

            if sheetEnv == "Instructions":
                # skip Instructions sheet 
                pass
            elif sheetEnv.strip().lower() == 'program details':
                print("--->Checking Program details sheet...")
                detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                      for
                                      col_index_env in range(detailsEnvSheet.ncols)}
                    programNameInp = dictDetailsEnv['Title of the Program'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Title of the Program'] else terminatingMessage("\"Title of the Program\" must not be Empty in \"Program details\" sheet")
                    extIdPGM = dictDetailsEnv['Program ID'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Program ID'] else terminatingMessage("\"Program ID\" must not be Empty in \"Program details\" sheet")
                    returnvalues = []
                    global entitiesPGM
                    entitiesPGM = dictDetailsEnv['Targeted state at program level'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Targeted state at program level'] else terminatingMessage("\"Targeted state at program level\" must not be Empty in \"Program details\" sheet")
                    districtentitiesPGM = dictDetailsEnv['Targeted district at program level'].encode('utf-8').decode('utf-8')
                    global startDateOfProgram, endDateOfProgram
                    startDateOfProgram = dictDetailsEnv['Start date of program']
                    endDateOfProgram = dictDetailsEnv['End date of program']

                    # taking the start date of program from program template and converting YYYY-MM-DD 00:00:00 format
                    
                    startDateArr = str(startDateOfProgram).split("-")
                    startDateOfProgram = startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[0] + " 00:00:00"

                    # taking the end date of program from program template and converting YYYY-MM-DD 00:00:00 format

                    endDateArr = str(endDateOfProgram).split("-")
                    endDateOfProgram = endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"
                        
                    global scopeEntityType
                    scopeEntityType = "state"


                    if districtentitiesPGM:
                        entitiesPGM = districtentitiesPGM
                        EntityType = "district"
                    else:
                        entitiesPGM = entitiesPGM
                        EntityType = "state"

                    scopeEntityType = EntityType

                    global entitiesPGMID
                    entitiesPGMID = fetchEntityId(parentFolder, accessToken,
                                                  entitiesPGM.lstrip().rstrip().split(","), scopeEntityType)
                    global orgIds
                    if environment == "staging":
                        orgIds = "01269934121990553633"
                    elif environment == "dev":
                        orgIds = "0137541424673095687"
                    else:
                        orgIds=fetchOrgId(environment, accessToken, parentFolder, OrgName)
                    # print(orgIds)
                    # sys.exit()

                    if not getProgramInfo(accessToken, parentFolder, programNameInp.encode('utf-8').decode('utf-8')):
                        extIdPGM = dictDetailsEnv['Program ID'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Program ID'] else terminatingMessage("\"Program ID\" must not be Empty in \"Program details\" sheet")
                        if str(dictDetailsEnv['Program ID']).strip() == "Do not fill this field":
                            terminatingMessage("change the program id")
                        descriptionPGM = dictDetailsEnv['Description of the Program'].encode('utf-8').decode('utf-8') if dictDetailsEnv[
                            'Description of the Program'] else terminatingMessage(
                            "\"Description of the Program\" must not be Empty in \"Program details\" sheet")
                        keywordsPGM = dictDetailsEnv['Keywords'].encode('utf-8').decode('utf-8')
                        entitiesPGM = dictDetailsEnv['Targeted state at program level'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Targeted state at program level'] else terminatingMessage("\"Targeted state at program level\" must not be Empty in \"Program details\" sheet")
                        districtentitiesPGM = dictDetailsEnv['Targeted district at program level'].encode('utf-8').decode('utf-8')
                        # selecting entity type based on the users input 
                        if districtentitiesPGM:
                            entitiesPGM = districtentitiesPGM
                            EntityType = "district"
                        else:
                            entitiesPGM = entitiesPGM
                            EntityType = "state"

                        scopeEntityType = EntityType

                        mainRole = dictDetailsEnv['Targeted role at program level'] if dictDetailsEnv['Targeted role at program level'] else terminatingMessage("\"Targeted role at program level\" must not be Empty in \"Program details\" sheet")
                        global rolesPGM
                        rolesPGM = dictDetailsEnv['Targeted subrole at program level'] if dictDetailsEnv['Targeted subrole at program level'] else terminatingMessage("\"Targeted subrole at program level\" must not be Empty in \"Program details\" sheet")
                        
                        if "teacher" in mainRole.strip().lower():
                            rolesPGM = str(rolesPGM).strip() + ",TEACHER"
                        if environment == "staging":
                            userDetails = ["5d7255bb-1216-460e-9228-59b60230b1c1","stagingpd_wjtv","Stagingpd",["PROGRAM_DESIGNER"],"",""]
                        elif environment == "dev":
                            userDetails = ["469dc732-04f3-42d9-9a85-30957a797acc","content","Contentreviewer",["PROGRAM_DESIGNER"],"",""]
                        else:
                            # fetch user details 
                            userDetails = fetchUserDetails(environment, accessToken, dictDetailsEnv['Diksha username/user id/email id/phone no. of Program Designer']).encode('utf-8').decode('utf-8')
                        creatorKeyCloakId = userDetails[0]
                        creatorName = userDetails[2]
                        
                        messageArr = []

                        scopeEntityType = EntityType
                        # fetch entity details 
                        entitiesPGMID = fetchEntityId(parentFolder, accessToken,entitiesPGM.lstrip().rstrip().split(","), scopeEntityType)
                        print(entitiesPGMID)
                        # sys.exit()
                        # fetch sub-role details 
                        rolesPGMID = fetchScopeRole(parentFolder, accessToken, rolesPGM.lstrip().rstrip().split(","))
                        print(rolesPGM)
                        # sys.exit()

                        # call function to create program 
                        programCreation(accessToken, parentFolder, extIdPGM, programNameInp, descriptionPGM,keywordsPGM.lstrip().rstrip().split(","), entitiesPGMID, rolesPGMID, orgIds,creatorKeyCloakId, creatorName,entitiesPGM,mainRole,rolesPGM)
                        # sys.exit()
                        programmappingpdpmsheetcreation(MainFilePath, accessToken, program_file, extIdPGM,parentFolder)

                        # map PM / PD to the program 
                        Programmappingapicall(MainFilePath, accessToken, program_file,parentFolder)

                        # check if program is created or not 
                        if getProgramInfo(accessToken, parentFolder, extIdPGM):
                            print("Program Created SuccessFully.")
                        else :
                            terminatingMessage("Program creation failed! Please check logs.")

            elif sheetEnv.strip().lower() == 'resource details':
                # checking Resource details sheet 
                print("--->Checking Resource Details sheet...")
                detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                           range(detailsEnvSheet.ncols)]
                # iterate through each row in Resource Details sheet and validate 
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                      for
                                      col_index_env in range(detailsEnvSheet.ncols)}
                    resourceNamePGM = dictDetailsEnv['Name of resources in program'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Name of resources in program'] else terminatingMessage("\"Name of resources in program\" must not be Empty in \"Resource Details\" sheet")
                    resourceTypePGM = dictDetailsEnv['Type of resources'].encode('utf-8').decode('utf-8') if dictDetailsEnv['Type of resources'] else terminatingMessage("\"Type of resources\" must not be Empty in \"Resource Details\" sheet")
                    resourceLinkOrExtPGM = dictDetailsEnv['Resource Link']
                    resourceStatusOrExtPGM = dictDetailsEnv['Resource Status'] if dictDetailsEnv['Resource Status'] else terminatingMessage("\"Resource Status\" must not be Empty in \"Resource Details\" sheet")
                    # setting start and end dates globally. 
                    global startDateOfResource, endDateOfResource
                    startDateOfResource = dictDetailsEnv['Start date of resource']
                    endDateOfResource = dictDetailsEnv['End date of resource']
                    # checking resource types and calling relevant functions 
                    if resourceTypePGM.lstrip().rstrip().lower() == "course":
                        coursemapping = courseMapToProgram(accessToken, resourceLinkOrExtPGM, parentFolder)
                        if startDateOfResource:
                            startDateArr = str(startDateOfResource).split("-")
                            bodySolutionUpdate = {"startDate": startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[0] + " 00:00:00"}
                            solutionUpdate(parentFolder, accessToken, coursemapping, bodySolutionUpdate)
                        if endDateOfResource:
                            endDateArr = str(endDateOfResource).split("-")
                            bodySolutionUpdate = {
                                "endDate": endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"}
                            solutionUpdate(parentFolder, accessToken, coursemapping, bodySolutionUpdate)


def fetchSolutionDetailsFromProgramSheet(solutionName_for_folder_path, programFile, solutionId, accessToken):
    global solutionRolesArray, solutionStartDate, solutionEndDate
    urlFetchSolutionApi = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'fetchSolutionDoc') + solutionId
    headerFetchSolutionApi = {
        'Authorization': config.get(environment, 'Authorization'),
        'X-authenticated-user-token': accessToken,
        'X-Channel-id': config.get(environment, 'X-Channel-id'),
        'internal-access-token': config.get(environment, 'internal-access-token')
    }
    payloadFetchSolutionApi = {}

    responseFetchSolutionApi = requests.post(url=urlFetchSolutionApi, headers=headerFetchSolutionApi,
                                             data=payloadFetchSolutionApi)
    responseFetchSolutionJson = responseFetchSolutionApi.json()
    messageArr = ["Solution Fetch Link.",
                  "solution name : " + responseFetchSolutionJson["result"]["name"],
                  "solution ExternalId : " + responseFetchSolutionJson["result"]["externalId"]]
    messageArr.append("Upload status code : " + str(responseFetchSolutionApi.status_code))
    createAPILog(solutionName_for_folder_path, messageArr)

    if responseFetchSolutionApi.status_code == 200:
        print('Fetch solution Api Success')
        
        solutionName = responseFetchSolutionJson["result"]["name"]

        xfile = openpyxl.load_workbook(programFile)

        resourceDetailsSheet = xfile.get_sheet_by_name('Resource Details')
        rowCountRD = resourceDetailsSheet.max_row
        columnCountRD = resourceDetailsSheet.max_column
        for row in range(3, rowCountRD + 1):
            if resourceDetailsSheet["A" + str(row)].value == solutionName:
                solutionMainRole = str(resourceDetailsSheet["E" + str(row)].value).strip()
                solutionRolesArray = str(resourceDetailsSheet["F" + str(row)].value).split(",") if str(
                    resourceDetailsSheet["E" + str(row)].value).split(",") else []
                if "teacher" in solutionMainRole.strip().lower():
                    solutionRolesArray.append("TEACHER")
                solutionStartDate = resourceDetailsSheet["G" + str(row)].value
                solutionEndDate = resourceDetailsSheet["H" + str(row)].value
    return [solutionRolesArray, solutionStartDate, solutionEndDate]

def prepareProgramSuccessSheet(MainFilePath, solutionName_for_folder_path, programFile, solutionExternalId, solutionId,accessToken):
    urlFetchSolutionApi = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'fetchSolutionDoc') + solutionId
    headerFetchSolutionApi = {
        'Authorization': config.get(environment, 'Authorization'),
        'X-authenticated-user-token': accessToken,
        'X-Channel-id': config.get(environment, 'X-Channel-id'),
        'internal-access-token': config.get(environment, 'internal-access-token')
    }
    payloadFetchSolutionApi = {}

    responseFetchSolutionApi = requests.post(url=urlFetchSolutionApi, headers=headerFetchSolutionApi,
                                             data=payloadFetchSolutionApi)
    responseFetchSolutionJson = responseFetchSolutionApi.json()
    messageArr = ["Solution Fetch Link.",
                  "solution name : " + responseFetchSolutionJson["result"]["name"],
                  "solution ExternalId : " + responseFetchSolutionJson["result"]["externalId"]]
    messageArr.append("Upload status code : " + str(responseFetchSolutionApi.status_code))
    createAPILog(solutionName_for_folder_path, messageArr)

    if responseFetchSolutionApi.status_code == 200:
        print('Fetch solution Api Success')
        solutionName = responseFetchSolutionJson["result"]["name"]
    urlFetchSolutionLinkApi = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'fetchLink') + solutionId
    headerFetchSolutionLinkApi = {
        'Authorization': config.get(environment, 'Authorization'),
        'X-authenticated-user-token': accessToken,
        'X-Channel-id': config.get(environment, 'X-Channel-id'),
        'internal-access-token': config.get(environment, 'internal-access-token')
    }
    payloadFetchSolutionLinkApi = {}

    responseFetchSolutionLinkApi = requests.post(url=urlFetchSolutionLinkApi, headers=headerFetchSolutionLinkApi,
                                                 data=payloadFetchSolutionLinkApi)
    messageArr = ["Solution Fetch Link.","solution id : " + solutionId,"solution ExternalId : " + solutionExternalId]
    messageArr.append("Upload status code : " + str(responseFetchSolutionLinkApi.status_code))
    createAPILog(solutionName_for_folder_path, messageArr)

    if responseFetchSolutionLinkApi.status_code == 200:
        print('Fetch solution Link Api Success')
        responseProjectUploadJson = responseFetchSolutionLinkApi.json()
        solutionLink = responseProjectUploadJson["result"]
        messageArr.append("Response : " + str(responseFetchSolutionLinkApi.text))
        createAPILog(solutionName_for_folder_path, messageArr)

        if os.path.exists(MainFilePath + "/" + str(programFile).replace(".xlsx", "") + '-SuccessSheet.xlsx'):
            xfile = openpyxl.load_workbook(
                MainFilePath + "/" + str(programFile).replace(".xlsx", "") + '-SuccessSheet.xlsx')
        else:
            xfile = openpyxl.load_workbook(programFile)

        resourceDetailsSheet = xfile.get_sheet_by_name('Resource Details')

        greenFill = PatternFill(start_color='0000FF00',
                                end_color='0000FF00',
                                fill_type='solid')
        rowCountRD = resourceDetailsSheet.max_row
        columnCountRD = resourceDetailsSheet.max_column
        for row in range(3, rowCountRD + 1):
            if str(resourceDetailsSheet["B" + str(row)].value).rstrip().lstrip().lower() == "course":
                resourceDetailsSheet["D1"] = ""
                resourceDetailsSheet["E1"] = ""
                resourceDetailsSheet['I2'] = "External id of the resource"
                resourceDetailsSheet['J2'] = "link to access the resource/Response"
                resourceDetailsSheet['I2'].fill = greenFill
                resourceDetailsSheet['J2'].fill = greenFill
                resourceDetailsSheet['I' + str(row)] = solutionExternalId
                resourceDetailsSheet['J' + str(row)] = "The course has been successfully mapped to the program"
                resourceDetailsSheet['I' + str(row)].fill = greenFill
                resourceDetailsSheet['J' + str(row)].fill = greenFill
            elif str(resourceDetailsSheet["A" + str(row)].value).strip() == solutionName:
                resourceDetailsSheet["D1"] = ""
                resourceDetailsSheet["E1"] = ""
                resourceDetailsSheet['I2'] = "External id of the resource"
                resourceDetailsSheet['J2'] = "link to access the resource/Response"
                resourceDetailsSheet['I2'].fill = greenFill
                resourceDetailsSheet['J2'].fill = greenFill
                resourceDetailsSheet['I' + str(row)] = solutionExternalId
                resourceDetailsSheet['J' + str(row)] = solutionLink
                resourceDetailsSheet['I' + str(row)].fill = greenFill
                resourceDetailsSheet['J' + str(row)].fill = greenFill

        programFile = str(programFile).replace(".xlsx", "")
        xfile.save(MainFilePath + "/" + programFile + '-SuccessSheet.xlsx')
        print("Program success sheet is created")

    else:
        print("Fetch solution link API Failed")
        messageArr.append("Response : " + str(responseFetchSolutionLinkApi.text))
        createAPILog(solutionName_for_folder_path, messageArr)
        sys.exit()


def solutionUpdate(solutionName_for_folder_path, accessToken, solutionId, bodySolutionUpdate):
    solutionUpdateApi = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, 'solutionUpdateApi') + str(solutionId)
    headerUpdateSolutionApi = {
        'Content-Type': 'application/json',
        'Authorization': config.get(environment, 'Authorization'),
        'X-authenticated-user-token': accessToken,
        'X-Channel-id': config.get(environment, 'X-Channel-id'),
        "internal-access-token": config.get(environment, 'internal-access-token')
        }
    responseUpdateSolutionApi = requests.post(url=solutionUpdateApi, headers=headerUpdateSolutionApi,data=json.dumps(bodySolutionUpdate))
    messageArr = ["Solution Update API called.", "URL : " + str(solutionUpdateApi), "Body : " + str(bodySolutionUpdate),"Response : " + str(responseUpdateSolutionApi.text),"Status Code : " + str(responseUpdateSolutionApi.status_code)]
    createAPILog(solutionName_for_folder_path, messageArr)
    if responseUpdateSolutionApi.status_code == 200:
        print("Solution Update Success.")
        return True
    else:
        print("Solution Update Failed.")
        return False


# upload survey questions 
def uploadSurveyQuestions(parentFolder, wbSurvey, addObservationSolution, accessToken, surveySolutionExternalId, surveyParentSolutionId,millisecond):
    sheetNam = wbSurvey.sheet_names()
    stDt = None
    enDt = None
    shCnt = 0
    for i in sheetNam:
        if i.strip().lower() == 'questions':
            sheetNam1 = wbSurvey.sheets()[shCnt]
        shCnt = shCnt + 1
    dataSort = [sheetNam1.row_values(i) for i in range(sheetNam1.nrows)]
    labels = dataSort[1]
    dataSort = dataSort[2:]
    dataSort.sort(key=lambda x: int(x[0]))
    openWorkBookSort1 = xl_copy(wbSurvey)
    sheet1 = openWorkBookSort1.add_sheet('questions_sequence_sorted')

    for idx, label in enumerate(labels):
        sheet1.write(0, idx, label)

    for idx_r, row in enumerate(dataSort):
        for idx_c, value in enumerate(row):
            sheet1.write(idx_r + 1, idx_c, value)
    newFileName = str(addObservationSolution)
    openWorkBookSort1.save(newFileName)
    openNewFile = xlrd.open_workbook(newFileName, on_demand=True)
    wbSurvey = openNewFile
    sheetNames = wbSurvey.sheet_names()
    for sheet2 in sheetNames:
        if sheet2.strip().lower() == 'questions_sequence_sorted':
            questionsList = []
            questionsSheet = wbSurvey.sheet_by_name(sheet2.lower())
            keys2 = [questionsSheet.cell(0, col_index2).value for col_index2 in
                     range(questionsSheet.ncols)]
            for row_index2 in range(1, questionsSheet.nrows):
                d2 = {keys2[col_index2]: questionsSheet.cell(row_index2, col_index2).value
                      for col_index2 in range(questionsSheet.ncols)}
                questionsList.append(d2)
            questionSeqByEcmArr = []
            quesSeqCnt = 1.0
            questionUploadFieldnames = []
            questionUploadFieldnames = ['solutionId', 'instanceParentQuestionId','hasAParentQuestion', 'parentQuestionOperator','parentQuestionValue', 'parentQuestionId','externalId', 'question0', 'question1', 'tip','hint', 'instanceIdentifier', 'responseType','dateFormat', 'autoCapture', 'validation','validationIsNumber', 'validationRegex','validationMax', 'validationMin', 'file','fileIsRequired', 'fileUploadType','allowAudioRecording', 'minFileCount','maxFileCount', 'caption', 'questionGroup','modeOfCollection', 'accessibility', 'showRemarks','rubricLevel', 'isAGeneralQuestion', 'R1','R1-hint', 'R2', 'R2-hint', 'R3', 'R3-hint', 'R4','R4-hint', 'R5', 'R5-hint', 'R6', 'R6-hint', 'R7','R7-hint', 'R8', 'R8-hint', 'R9', 'R9-hint', 'R10','R10-hint', 'R11', 'R11-hint', 'R12', 'R12-hint','R13', 'R13-hint', 'R14', 'R14-hint', 'R15','R15-hint', 'R16', 'R16-hint', 'R17', 'R17-hint','R18', 'R18-hint', 'R19', 'R19-hint', 'R20','R20-hint', 'sectionHeader', 'page','questionNumber', '_arrayFields']

            for ques in questionsList:

                questionFilePath = parentFolder + '/questionUpload/'
                file_exists_ques = os.path.isfile(
                    parentFolder + '/questionUpload/uploadSheet.csv')
                if not os.path.exists(questionFilePath):
                    os.mkdir(questionFilePath)
                with open(parentFolder + '/questionUpload/uploadSheet.csv', 'a',
                          encoding='utf-8') as questionUploadFile:
                    writerQuestionUpload = csv.DictWriter(questionUploadFile, fieldnames=questionUploadFieldnames)
                    if not file_exists_ques:
                        writerQuestionUpload.writeheader()
                    questionFileObj = {}
                    surveyExternalId = None
                    questionFileObj['solutionId'] = surveySolutionExternalId
                    if ques['instance_parent_question_id'].encode('utf-8').decode('utf-8'):
                        questionFileObj['instanceParentQuestionId'] = ques[
                                                                          'instance_parent_question_id'].strip() + '_' + str(
                            millisecond)
                    else:
                        questionFileObj['instanceParentQuestionId'] = 'NA'
                    if ques['parent_question_id'].encode('utf-8').decode('utf-8').strip():
                        questionFileObj['hasAParentQuestion'] = 'YES'
                        if ques['show_when_parent_question_value_is'] == 'or':
                            questionFileObj['parentQuestionOperator'] = '||'
                        else:
                            questionFileObj['parentQuestionOperator'] = ques['show_when_parent_question_value_is']
                        if type(ques['parent_question_value']) != str:
                            if (ques['parent_question_value'] and ques[
                                'parent_question_value'].is_integer() == True):
                                questionFileObj['parentQuestionValue'] = int(ques['parent_question_value'])
                            elif (ques['parent_question_value'] and ques[
                                'parent_question_value'].is_integer() == False):
                                questionFileObj['parentQuestionValue'] = ques['parent_question_value']
                        else:
                            questionFileObj['parentQuestionValue'] = ques['parent_question_value']
                            questionFileObj['parentQuestionId'] = ques['parent_question_id'].encode('utf-8').decode('utf-8').strip() + '_' + str(
                                millisecond)
                    else:
                        questionFileObj['hasAParentQuestion'] = 'NO'
                        questionFileObj['parentQuestionOperator'] = None
                        questionFileObj['parentQuestionValue'] = None
                        questionFileObj['parentQuestionId'] = None
                    questionFileObj['externalId'] = ques['question_id'].strip() + '_' + str(millisecond)
                    if quesSeqCnt == ques['question_sequence']:
                        questionSeqByEcmArr.append(ques['question_id'].strip() + '_' + str(millisecond))
                        quesSeqCnt = quesSeqCnt + 1.0
                    if ques['question_language1']:
                        questionFileObj['question0'] = ques['question_language1']
                    else:
                        questionFileObj['question0'] = None
                    if ques['question_language2']:
                        questionFileObj['question1'] = ques['question_language2'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['question1'] = None
                    if ques['question_tip']:
                        questionFileObj['tip'] = ques['question_tip'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['tip'] = None
                    if ques['question_hint']:
                        questionFileObj['hint'] = ques['question_hint'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['hint'] = None
                    if ques['instance_identifier']:
                        questionFileObj['instanceIdentifier'] = ques['instance_identifier'].encode('utf-8').decode('utf-8')
                    else:
                        questionFileObj['instanceIdentifier'] = None
                    if ques['question_response_type'].strip().lower():
                        questionFileObj['responseType'] = ques['question_response_type'].strip().lower()
                    if ques['question_response_type'].strip().lower() == 'date':
                        questionFileObj['dateFormat'] = "DD-MM-YYYY"
                    else:
                        questionFileObj['dateFormat'] = None
                    if ques['question_response_type'].strip().lower() == 'date':
                        if ques['date_auto_capture'] and ques['date_auto_capture'] == 1:
                            questionFileObj['autoCapture'] = 'TRUE'
                        elif ques['date_auto_capture'] and ques['date_auto_capture'] == 0:
                            questionFileObj['autoCapture'] = 'false'
                        else:
                            questionFileObj['autoCapture'] = 'false'
                    else:
                        questionFileObj['autoCapture'] = None
                    if ques['response_required']:
                        if ques['response_required'] == 1:
                            questionFileObj['validation'] = 'TRUE'
                        elif ques['response_required'] == 0:
                            questionFileObj['validation'] = 'FALSE'
                    else:
                        questionFileObj['validation'] = 'FALSE'
                    if ques['question_response_type'].strip().lower() == 'number':
                        questionFileObj['validationIsNumber'] = 'TRUE'
                        questionFileObj['validationRegex'] = 'isNumber'
                        if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                            questionFileObj['validationMax'] = int(ques['max_number_value'])
                        elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                            questionFileObj['validationMax'] = ques['max_number_value']
                        else:
                            questionFileObj['validationMax'] = 10000

                        if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                            questionFileObj['validationMin'] = int(ques['min_number_value'])
                        elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                            questionFileObj['validationMin'] = ques['min_number_value']
                        else:
                            questionFileObj['validationMax'] = 10000

                        if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                            questionFileObj['validationMin'] = int(ques['min_number_value'])
                        elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                            questionFileObj['validationMin'] = ques['min_number_value']
                        else:
                            questionFileObj['validationMin'] = 0

                    elif ques['question_response_type'].strip().lower() == 'slider':
                        questionFileObj['validationIsNumber'] = None
                        questionFileObj['validationRegex'] = 'isNumber'
                        if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                            questionFileObj['validationMax'] = int(ques['max_number_value'])
                        elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                            questionFileObj['validationMax'] = ques['max_number_value']
                        else:
                            questionFileObj['validationMax'] = 5

                        if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                            questionFileObj['validationMin'] = int(ques['min_number_value'])
                        elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                            questionFileObj['validationMin'] = ques['min_number_value']
                        else:
                            questionFileObj['validationMin'] = 0
                    else:
                        questionFileObj['validationIsNumber'] = None
                        questionFileObj['validationRegex'] = None
                        questionFileObj['validationMax'] = None
                        questionFileObj['validationMin'] = None
                    if ques['file_upload'] == 1:
                        questionFileObj['file'] = 'Snapshot'
                        questionFileObj['fileIsRequired'] = 'TRUE'
                        questionFileObj['fileUploadType'] = 'image/jpeg,docx,pdf,ppt'
                        questionFileObj['minFileCount'] = 0
                        questionFileObj['maxFileCount'] = 10
                    elif ques['file_upload'] == 0:
                        questionFileObj['file'] = 'NA'
                        questionFileObj['fileIsRequired'] = None
                        questionFileObj['fileUploadType'] = None
                        questionFileObj['minFileCount'] = None
                        questionFileObj['maxFileCount'] = None

                    questionFileObj['caption'] = 'FALSE'
                    questionFileObj['questionGroup'] = 'A1'
                    questionFileObj['modeOfCollection'] = 'onfield'
                    questionFileObj['accessibility'] = 'No'
                    if ques['show_remarks'] == 1:
                        questionFileObj['showRemarks'] = 'TRUE'
                    elif ques['show_remarks'] == 0:
                        questionFileObj['showRemarks'] = 'FALSE'
                    questionFileObj['rubricLevel'] = None
                    questionFileObj['isAGeneralQuestion'] = None
                    if ques['question_response_type'].strip().lower() == 'radio' or ques[
                        'question_response_type'].strip() == 'multiselect':
                        for quesIndex in range(1, 21):
                            if type(ques['response(R' + str(quesIndex) + ')']) != str:
                                if (ques['response(R' + str(quesIndex) + ')'] and ques[
                                    'response(R' + str(quesIndex) + ')'].is_integer() == True):
                                    questionFileObj['R' + str(quesIndex) + ''] = int(
                                        ques['response(R' + str(quesIndex) + ')'])
                                elif (ques['response(R' + str(quesIndex) + ')'] and ques[
                                    'response(R' + str(quesIndex) + ')'].is_integer() == False):
                                    questionFileObj['R' + str(quesIndex) + ''] = ques[
                                        'response(R' + str(quesIndex) + ')']
                            else:
                                questionFileObj['R' + str(quesIndex) + ''] = ques[
                                    'response(R' + str(quesIndex) + ')']

                            if type(ques['response(R' + str(quesIndex) + ')_hint']) != str:
                                if (ques['response(R' + str(quesIndex) + ')_hint'] and ques[
                                    'response(R' + str(quesIndex) + ')_hint'].is_integer() == True):
                                    questionFileObj['R' + str(quesIndex) + '-hint'] = int(
                                        ques['response(R' + str(quesIndex) + ')_hint'])
                                elif (ques['response(R' + str(quesIndex) + ')_hint'] and ques[
                                    'response(R' + str(quesIndex) + ')_hint'].is_integer() == False):
                                    questionFileObj['R' + str(quesIndex) + '-hint'] = ques[
                                        'response(R' + str(quesIndex) + ')_hint']
                            else:
                                questionFileObj['R' + str(quesIndex) + '-hint'] = ques[
                                    'response(R' + str(quesIndex) + ')_hint']
                            questionFileObj['_arrayFields'] = 'parentQuestionValue'
                    else:
                        for quesIndex in range(1, 21):
                            questionFileObj['R' + str(quesIndex)] = None
                            questionFileObj['R' + str(quesIndex) + '-hint'] = None
                    if ques['section_header'].encode('utf-8').decode('utf-8'):
                        questionFileObj['sectionHeader'] = ques['section_header']
                    else:
                        questionFileObj['sectionHeader'] = None

                    questionFileObj['page'] = ques['page']
                    if type(ques['question_number']) != str:
                        if ques['question_number'] and ques['question_number'].is_integer() == True:
                            questionFileObj['questionNumber'] = int(ques['question_number'])
                        elif ques['question_number']:
                            questionFileObj['questionNumber'] = ques['question_number']
                        else:
                            questionFileObj['questionNumber'] = ques['question_number']
                    writerQuestionUpload.writerow(questionFileObj)
                    # print(questionFileObj)
                # terminatingMessage("Question")
            urlQuestionsUploadApi = config.get(environment, 'INTERNAL_KONG_IP')+ config.get(environment, 'questionUploadApiUrl')
            headerQuestionUploadApi = {
                'Authorization': config.get(environment, 'Authorization'),
                'X-authenticated-user-token': accessToken,
                'X-Channel-id': config.get(environment, 'X-Channel-id')
            }
            filesQuestion = {
                'questions': open(parentFolder + '/questionUpload/uploadSheet.csv', 'rb')
            }
            responseQuestionUploadApi = requests.post(url=urlQuestionsUploadApi,
                                                      headers=headerQuestionUploadApi, files=filesQuestion)
            if responseQuestionUploadApi.status_code == 200:
                print('Question upload Success')

                messageArr = ["********* Question Upload api *********", "URL : " + urlQuestionsUploadApi,
                              "Path : " + str(parentFolder) + str('/questionUpload/uploadSheet.csv'),
                              "Status code : " + str(responseQuestionUploadApi.status_code),
                              "Response : " + responseQuestionUploadApi.text]
                createAPILog(parentFolder, messageArr)
                messageArr1 = ["Questions","Question upload Success","Passed",str(responseQuestionUploadApi.status_code)]
                apicheckslog(parentFolder,messageArr1)

                with open(parentFolder + '/questionUpload/uploadInternalIdsSheet.csv', 'w+',encoding='utf-8') as questionRes:
                    questionRes.write(responseQuestionUploadApi.text)
                urlImportSoluTemplate = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment,'importSurveySolutionTemplateUrl') + str(surveyParentSolutionId) + "?appName=manage-learn"
                headerImportSoluTemplateApi = {
                    'Authorization': config.get(environment, 'Authorization'),
                    'X-authenticated-user-token': accessToken,
                    'X-Channel-id': config.get(environment, 'X-Channel-id')
                }
                responseImportSoluTemplateApi = requests.get(url=urlImportSoluTemplate,
                                                             headers=headerImportSoluTemplateApi)
                if responseImportSoluTemplateApi.status_code == 200:
                    print('Creating Child Success')

                    messageArr = ["********* Creating Child api *********", "URL : " + urlImportSoluTemplate,
                                  "Status code : " + str(responseImportSoluTemplateApi.status_code),
                                  "Response : " + responseImportSoluTemplateApi.text]
                    createAPILog(parentFolder, messageArr)
                    responseImportSoluTemplateApi = responseImportSoluTemplateApi.json()
                    solutionIdSuc = responseImportSoluTemplateApi["result"]["solutionId"]
                    urlSurveyProgramMapping = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment, "importSurveySolutionToProgramUrl") + str(solutionIdSuc) + "?programId=" + programExternalId.lstrip().rstrip()
                    headeSurveyProgramMappingApi = {
                        'Authorization': config.get(environment, 'Authorization'),
                        'X-authenticated-user-token': accessToken,
                        'X-Channel-id': config.get(environment, 'X-Channel-id')
                    }
                    responseSurveyProgramMappingApi = requests.get(url=urlSurveyProgramMapping,headers=headeSurveyProgramMappingApi)
                    if responseSurveyProgramMappingApi.status_code == 200:
                        print('Program Mapping Success')
                        
                        messageArr = ["********* Program mapping api *********", "URL : " + urlSurveyProgramMapping,
                                      "Status code : " + str(responseSurveyProgramMappingApi.status_code),
                                      "Response : " + responseSurveyProgramMappingApi.text]
                        createAPILog(parentFolder, messageArr)
                        surveyLink = None
                        solutionIdSuc = None
                        surveyExternalIdSuc = None
                        surveyLink = responseImportSoluTemplateApi["result"]["link"]
                        solutionIdSuc = responseImportSoluTemplateApi["result"]["solutionId"]
                        solutionExtIdSuc = responseImportSoluTemplateApi["result"]["solutionExternalId"]
                        print("Survey Child Id : " + str(solutionExtIdSuc))
                        solutionDetails = fetchSolutionDetailsFromProgramSheet(parentFolder, programFile, solutionIdSuc,
                                                                               accessToken)
                        scopeEntities = entitiesPGMID
                        scopeRoles = solutionDetails[0]
                        surveyScopeBody = {
                            "scope": {"entityType": scopeEntityType, "entities": scopeEntities, "roles": scopeRoles}}
                        solutionUpdate(parentFolder, accessToken, solutionIdSuc, surveyScopeBody)
                        prepareProgramSuccessSheet(MainFilePath, parentFolder, programFile, solutionExtIdSuc,
                                                   solutionIdSuc, accessToken)
                        
                        print('Survey Successfully Added')
                    else:
                        print('Program Mapping Failed')
                        messageArr = ["********* Program mapping api *********", "URL : " + urlSurveyProgramMapping,
                                      "Status code : " + str(responseSurveyProgramMappingApi.status_code),
                                      "Response : " + responseSurveyProgramMappingApi.text]
                        createAPILog(parentFolder, messageArr)
                else:
                    print('Creating Child API Failed')
                    messageArr = ["********* Program mapping api *********", "URL : " + urlImportSoluTemplate,
                                  "Status code : " + str(responseImportSoluTemplateApi.status_code),
                                  "Response : " + responseImportSoluTemplateApi.text]
                    createAPILog(parentFolder, messageArr)
            else:
                print('QuestionUploadApi Failed')
                messageArr = ["********* Question Upload api *********", "URL : " + urlQuestionsUploadApi,
                              "Path : " + str(parentFolder) + str('/questionUpload/uploadSheet.csv'),
                              "Status code : " + str(responseQuestionUploadApi.status_code),
                              "Response : " + responseQuestionUploadApi.text]
                createAPILog(parentFolder, messageArr)


def checkEmailValidation(email):
    if (re.search(regex, email)):
        return True
    else:
        return False

# Fetch user details 
def fetchUserDetails(environment, accessToken, dikshaId):
    url = config.get(environment, 'host') + config.get(environment, 'userInfoApiUrl')
    messageArr = ["User search API called."]
    headers = {'Content-Type': 'application/json',
               'Authorization': config.get(environment, 'AuthorizationForHost'),
               'x-authenticated-user-token': accessToken}
    isEmail = checkEmailValidation(dikshaId.lstrip().rstrip())
    if isEmail:
        body = "{\n  \"request\": {\n    \"filters\": {\n    \t\"email\": \"" + dikshaId.lstrip().rstrip() + "\"\n    },\n      \"fields\" :[],\n    \"limit\": 1000,\n    \"sort_by\": {\"createdDate\": \"desc\"}\n  }\n}"
    else:
        body = "{\n  \"request\": {\n    \"filters\": {\n    \t\"userName\": \"" + dikshaId.lstrip().rstrip() + "\"\n    },\n      \"fields\" :[],\n    \"limit\": 1000,\n    \"sort_by\": {\"createdDate\": \"desc\"}\n  }\n}"

    responseUserSearch = requests.request("POST", url, headers=headers, data=body)
    if responseUserSearch.status_code == 200:
        responseUserSearch = responseUserSearch.json()
        if responseUserSearch['result']['response']['content']:
            userKeycloak = responseUserSearch['result']['response']['content'][0]['userId']
            userName = responseUserSearch['result']['response']['content'][0]['userName']
            firstName = responseUserSearch['result']['response']['content'][0]['firstName']
            rootOrgId = responseUserSearch['result']['response']['content'][0]['rootOrgId']
            for index in responseUserSearch['result']['response']['content'][0]['organisations']:
                if rootOrgId == index['organisationId']:
                    roledetails = index['roles']
                    rootOrgName = index['orgName']
            print(roledetails)
        else:
            terminatingMessage("-->Given username/email is not present in DIKSHA platform<--.")
    else:
        print(responseUserSearch.text)
        terminatingMessage("User fetch API failed. Check logs.")
    return [userKeycloak, userName, firstName,roledetails,rootOrgName,rootOrgId]

def fetchOrgId(environment, accessToken, parentFolder, OrgName):
    url = config.get(environment, 'host') + config.get(environment, 'fetchOrgDetails')
    messageArr = ["Org search API called."]
    headers = {'Content-Type': 'application/json',
               'Authorization': config.get(environment, 'Authorization'),
               'x-authenticated-user-token': accessToken}
    orgIds = []
    organisations = str(OrgName).split(",")
    for org in organisations:
        orgBody = {"id": "",
                   "ts": "",
                   "params": {
                       "msgid": "",
                       "resmsgid": "",
                       "status": "success"
                   },
                   "request": {
                       "filters": {
                           "orgName": str(org).strip()
                       }
                   }}

        responseOrgSearch = requests.request("POST", url, headers=headers, data=json.dumps(orgBody))
        if responseOrgSearch.status_code == 200:
            responseOrgSearch = responseOrgSearch.json()
            if responseOrgSearch['result']['response']['content']:
                orgId = responseOrgSearch['result']['response']['content'][0]['id']
                orgIds.append(orgId)
                messageArr.append("orgApi : " + str(url))
                messageArr.append("orgBody : " + str(orgBody))
                messageArr.append("orgAPI response: " + str(responseOrgSearch))
                messageArr.append("orgIds : " + str(orgIds))
            elif environment == "staging":
                messageArr.append("Given Organisation/ State tenant is not present in DIKSHA platform.")
                print("Given Organisation/ State tenant is not present in DIKSHA platform.")
                messageArr.append("orgApi : " + str(url))
                messageArr.append("orgBody : " + str(orgBody))
                messageArr.append("orgAPI response: " + str(responseOrgSearch))
            else:
                terminatingMessage("Given Organisation/ State tenant is not present in DIKSHA platform.")
                messageArr.append("orgApi : " + str(url))
                messageArr.append("orgBody : " + str(orgBody))
                messageArr.append("orgAPI response: " + str(responseOrgSearch))
        else:
            messageArr.append("orgApi : " + str(url))
            messageArr.append("headers : " + str(headers))
            messageArr.append("orgBody : " + str(orgBody))
            # print(str(orgBody))
            createAPILog(parentFolder, messageArr)
            print(responseOrgSearch.text)
            terminatingMessage("Organisation/ State tenant fetch API failed. Check logs.")
    return orgIds





def createChild(solutionName_for_folder_path, observationExternalId, accessToken):
    childObservationExternalId = str(observationExternalId + "_CHILD")
    urlSol_prog_mapping = config.get(environment,'solutionToprogramMAppingApiUrl') + "?solutionId=" + observationExternalId + "&entityType=" + entityType
    
    payloadSol_prog_mapping = {
        "externalId": childObservationExternalId,
        "name": solutionName.lstrip().rstrip(),
        "description": solutionDescription.lstrip().rstrip(),
        "programExternalId": programExternalId
    }
    headersSol_prog_mapping = {'Authorization': config.get(environment, 'Authorization'),
                               'X-authenticated-user-token': accessToken,
                               'Content-Type': config.get(environment, 'Content-Type')}
    responseSol_prog_mapping = requests.request("POST", urlSol_prog_mapping, headers=headersSol_prog_mapping,
                                                data=json.dumps(payloadSol_prog_mapping))
    messageArr = ["Create child API called.", "URL : " + urlSol_prog_mapping,
                  "Status code : " + str(responseSol_prog_mapping.status_code),
                  "Response : " + responseSol_prog_mapping.text, "body : " + str(payloadSol_prog_mapping)]
    if responseSol_prog_mapping.status_code == 200:
        print("Solution mapped to program : " + programName)
        print("Child solution : " + childObservationExternalId)

        responseSol_prog_mapping = responseSol_prog_mapping.json()
        child_id = responseSol_prog_mapping['result']['_id']
        createAPILog(solutionName_for_folder_path, messageArr)
        return [child_id, childObservationExternalId]
    else:
        print("Unable to create child solution")

        messageArr.append("Unable to create child solution")
        createAPILog(solutionName_for_folder_path, messageArr)
        return False

def createSurveySolution(parentFolder, wbSurvey, accessToken):
    sheetNames1 = wbSurvey.sheet_names()
    for sheetEnv in sheetNames1:
        if sheetEnv.strip().lower() == 'details':
            surveySolutionCreationReqBody = {}
            detailsEnvSheet = wbSurvey.sheet_by_name(sheetEnv)
            keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                       range(detailsEnvSheet.ncols)]

            for row_index_env in range(2, detailsEnvSheet.nrows):
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                  for
                                  col_index_env in range(detailsEnvSheet.ncols)}
                surveySolutionCreationReqBody['name'] = dictDetailsEnv['survey_solution_name'].encode('utf-8').decode('utf-8')
                surveySolutionCreationReqBody["description"] = dictDetailsEnv['survey_solution_description'].encode('utf-8').decode('utf-8')
                surveySolutionExternalId = str(uuid.uuid1())
                surveySolutionCreationReqBody["externalId"] = surveySolutionExternalId
                if dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8') == "":
                    exceptionHandlingFlag = True
                    print('survey_creator_username column should not be empty in the details sheet')
                    sys.exit()
                else:
                    surveySolutionCreationReqBody['creator'] = dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8')


                if environment == "staging":
                    userDetails = ["4cd4c690-eab6-4938-855a-447c7b1b8ea9","content_creator_tn3941","Harish"]
                elif environment == "dev":
                    userDetails = ["469dc732-04f3-42d9-9a85-30957a797acc","content","Contentreviewer"]
                else:    
                    userDetails = fetchUserDetails(environment, accessToken, dictDetailsEnv['survey_creator_username'])
                surveySolutionCreationReqBody['author'] = userDetails[0]

                # Below script will convert date DD-MM-YYYY TO YYYY-MM-DD 00:00:00 to match the code syntax 

                if dictDetailsEnv["survey_start_date"]:
                    if type(dictDetailsEnv["survey_start_date"]) == str:
                        startDateArr = None
                        startDateArr = (dictDetailsEnv["survey_start_date"]).split("-")
                        surveySolutionCreationReqBody["startDate"] = startDateArr[2] + "-" + startDateArr[1] + "-" + \
                                                                     startDateArr[0] + " 00:00:00"
                    elif type(dictDetailsEnv["survey_start_date"]) == float:
                        surveySolutionCreationReqBody["startDate"] = (
                            xlrd.xldate.xldate_as_datetime(dictDetailsEnv["survey_start_date"],
                                                           wbSurvey.datemode)).strftime("%Y/%m/%d")
                    else:
                        surveySolutionCreationReqBody["startDate"] = ""
                    if dictDetailsEnv["survey_end_date"]:
                        if type(dictDetailsEnv["survey_end_date"]) == str:
                            endDateArr = None
                            endDateArr = (dictDetailsEnv["survey_end_date"]).split("-")
                            surveySolutionCreationReqBody["endDate"] = endDateArr[2] + "-" + endDateArr[1] + "-" + \
                                                                       endDateArr[0] + " 23:59:59"
                        elif type(dictDetailsEnv["survey_end_date"]) == float:
                            surveySolutionCreationReqBody["endDate"] = (
                                xlrd.xldate.xldate_as_datetime(dictDetailsEnv["survey_end_date"],
                                                               wbSurvey.datemode)).strftime("%Y/%m/%d")
                        else:
                            surveySolutionCreationReqBody["endDate"] = ""
                        enDt = surveySolutionCreationReqBody["endDate"]
                        
                        urlCreateSolutionApi = config.get(environment, 'INTERNAL_KONG_IP')+ config.get(environment, 'surveySolutionCreationApiUrl')
                        headerCreateSolutionApi = {
                            'Content-Type': config.get(environment, 'Content-Type'),
                            'Authorization': config.get(environment, 'Authorization'),
                            'X-authenticated-user-token': accessToken,
                            'X-Channel-id': config.get(environment, 'X-Channel-id'),
                            'appName': config.get(environment, 'appName')
                        }
                        responseCreateSolutionApi = requests.post(url=urlCreateSolutionApi,
                                                                  headers=headerCreateSolutionApi,
                                                                  data=json.dumps(surveySolutionCreationReqBody))
                        responseInText = responseCreateSolutionApi.text
                        messageArr = ["********* Create Survey Solution *********", "URL : " + urlCreateSolutionApi,
                                      "BODY : " + str(surveySolutionCreationReqBody),
                                      "Status code : " + str(responseCreateSolutionApi.status_code),
                                      "Response : " + responseCreateSolutionApi.text]
                        fileheader = [(surveySolutionCreationReqBody['name']),('Program Sheet Validation'), ('noobra')]
                        createAPILog(parentFolder, messageArr)
                        apicheckslog(parentFolder,fileheader)
                        if responseCreateSolutionApi.status_code == 200:
                            responseCreateSolutionApi = responseCreateSolutionApi.json()
                            urlSearchSolution = config.get(environment, 'INTERNAL_KONG_IP') + config.get(environment,'fetchSolutionDetails') + "survey&page=1&limit=10&search=" + str(surveySolutionExternalId)
                            responseSearchSolution = requests.request("POST", urlSearchSolution,
                                                                      headers=headerCreateSolutionApi)
                            messageArr = ["********* Search Survey Solution *********", "URL : " + urlSearchSolution,
                                          "Status code : " + str(responseSearchSolution.status_code),
                                          "Response : " + responseSearchSolution.text]
                            createAPILog(parentFolder, messageArr)
                            # apicheckslog(parentFolder, messageArr)
                            if responseSearchSolution.status_code == 200:
                                responseSearchSolutionApi = responseSearchSolution.json()
                                surveySolutionExternalId = None
                                surveySolutionExternalId = responseSearchSolutionApi['result']['data'][0]['externalId']
                            else:
                                print("Solution fetch API failed")
                                print("URL : " + urlSearchSolution)
                                terminatingMessage("Status Code : " + responseSearchSolution.status_code)

                            solutionId = None
                            solutionId = responseCreateSolutionApi["result"]["solutionId"]
                            bodySolutionUpdate = {"creator": dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8')}
                            solutionUpdate(parentFolder, accessToken, solutionId, bodySolutionUpdate)

                            return [solutionId, surveySolutionExternalId]
                        else:
                            terminatingMessage("Survey creation Failed, check logs!")
    



# function to accept only csv file as input in command line argument
def valid_file(param):
    base, ext = os.path.splitext(param)
    if ext.lower() not in ('.xlsx'):
        raise argparse.ArgumentTypeError('File must have a csv extension')
    return param


#main execution
start_time = time.time()
parser = argparse.ArgumentParser()
parser.add_argument('--programFile', '--programFile', type=valid_file)
parser.add_argument('--env', '--env')
argument = parser.parse_args()
programFile = argument.programFile
environment = argument.env
millisecond = int(time.time() * 1000)

# Function create File structure for Solutions
def createFileStruct(MainFilePath, addObservationSolution):
    if not os.path.isdir(MainFilePath + '/SolutionFiles'):
        os.mkdir(MainFilePath + '/SolutionFiles')
    if "\\" in str(addObservationSolution):
        fileNameSplit = str(addObservationSolution).split('\\')[-1:]
    elif "/" in str(addObservationSolution):
        fileNameSplit = str(addObservationSolution).split('/')[-1:]
    else:
        fileNameSplit = str(addObservationSolution)
    if ".xlsx" in str(fileNameSplit[0]):
        ts = str(time.time()).replace(".", "_")
        folderName = fileNameSplit[0].replace(".xlsx", "-" + str(ts))
        os.mkdir(MainFilePath + '/SolutionFiles/' + str(folderName))
        path = os.path.join(MainFilePath + '/SolutionFiles', str(folderName))
        path = os.path.join(path, str('apiHitLogs'))
        os.mkdir(path)
    else:
        terminatingMessage("File Error.offff")
    returnPathStr = os.path.join(MainFilePath + '/SolutionFiles', str(folderName))

    if not os.path.isdir(returnPathStr + "/user_input_file"):
        os.mkdir(returnPathStr + "/user_input_file")

    shutil.copy(addObservationSolution, os.path.join(returnPathStr + "/user_input_file"))
    shutil.copy(programFile, os.path.join(returnPathStr + "/user_input_file"))
    return returnPathStr

# Function create File structure for Program
def createFileStructForProgram(programFile):
    if not os.path.isdir('programFiles'):
        os.mkdir('programFiles')
    if "\\" in str(programFile):
        fileNameSplit = str(programFile).split('\\')[-1:]
    elif "/" in str(programFile):
        fileNameSplit = str(programFile).split('/')[-1:]
    else:
        fileNameSplit = str(programFile)
    if ".xlsx" in fileNameSplit:
        ts = str(time.time()).replace(".", "_")
        folderName = fileNameSplit.replace(".xlsx", "-" + str(ts))
        os.mkdir('programFiles/' + str(folderName))
        path = os.path.join('programFiles', str(folderName))
    else:
        terminatingMessage("File Error.")
    returnPathStr = os.path.join('programFiles', str(folderName))

    return returnPathStr

# function to check environment 
def envCheck():
    try:
        config.get(environment, 'keyclockAPIUrl')
        return True
    except Exception as e:
        print(e)
        return False
    
def mainFunc(MainFilePath, programFile, addObservationSolution, millisecond, isProgramnamePresent, isCourse,
             scopeEntityType=scopeEntityType):
    scopeEntityType = scopeEntityType

    if not isCourse:
        parentFolder = createFileStruct(MainFilePath, addObservationSolution)
        accessToken = generateAccessToken(parentFolder)
        programsFileCheck(programFile, accessToken, parentFolder, MainFilePath)
        typeofSolution = validateSheets(addObservationSolution, accessToken, parentFolder)
        # sys.exit()
        wbObservation = xlrd.open_workbook(addObservationSolution, on_demand=True)
        wbProgram = xlrd.open_workbook(programFile, on_demand=True)
        # if typeofSolution == 3:
        surveyResp = createSurveySolution(parentFolder, wbObservation, accessToken)
        surTempExtID = surveyResp[1]
        bodySolutionUpdate = {"status": "active", "isDeleted": False}
        solutionUpdate(parentFolder, accessToken, surveyResp[0], bodySolutionUpdate)
        uploadSurveyQuestions(parentFolder, wbObservation, addObservationSolution, accessToken, surTempExtID,
                                  surveyResp[0], millisecond)                                      
# Generate access token for the APIs. 
def generateAccessToken(solutionName_for_folder_path):
    print("something")
    headerKeyClockUser = {'Content-Type': config.get(environment, 'keyclockAPIContent-Type')}
    
    responseKeyClockUser = requests.post(url=config.get(environment, 'host') + config.get(environment, 'keyclockAPIUrl'), headers=headerKeyClockUser,
                                         data=config.get(environment, 'keyclockAPIBody'))
    print(responseKeyClockUser)
    messageArr = []
    messageArr.append("URL : " + str(config.get(environment, 'keyclockAPIUrl')))
    messageArr.append("Body : " + str(config.get(environment, 'keyclockAPIBody')))
    messageArr.append("Status Code : " + str(responseKeyClockUser.status_code))
    if responseKeyClockUser.status_code == 200:
        responseKeyClockUser = responseKeyClockUser.json()
        accessTokenUser = responseKeyClockUser['access_token']
        messageArr.append("Acccess Token : " + str(accessTokenUser))
        createAPILog(solutionName_for_folder_path, messageArr)
        fileheader = ["Access Token","Access Token succesfully genarated","Passed"]
        apicheckslog(solutionName_for_folder_path,fileheader)
        print("--->Access Token Generated!")
        return accessTokenUser
    print("Error in generating Access token")
    print("Status code : " + str(responseKeyClockUser.status_code))
    createAPILog(solutionName_for_folder_path, messageArr)
    fileheader = ["Access Token", "Error in generating Access token", "Failed",responseKeyClockUser.status_code+"Check access token api"]
    apicheckslog(solutionName_for_folder_path, fileheader)
    fileheader = ["Access Token", "Error in generating Access token", "Failed","Check Headers of api"]
    apicheckslog(solutionName_for_folder_path, fileheader)
    terminatingMessage("Please check API logs.")

if envCheck():
    print("=================== Environment set to " + str(environment) + "=====================")
else:
    terminatingMessage(str(environment) + " is an invalid environment")
MainFilePath = createFileStructForProgram(programFile)
wbPgm = xlrd.open_workbook(programFile, on_demand=True)
sheetNames = wbPgm.sheet_names()
pgmSheets = ["Instructions", "Program Details", "Resource Details","Program Manager Details"]
print(sheetNames)
print(pgmSheets)
if len(sheetNames) == len(pgmSheets) and sheetNames == pgmSheets:
    print("--->Program Template detected.<---")
    for sheetEnv in sheetNames:
        if sheetEnv.strip().lower() == 'program details':
            print("Checking program details sheet...")
            programDetailsSheet = wbPgm.sheet_by_name(sheetEnv)
            keysEnv = [programDetailsSheet.cell(1, col_index_env).value for col_index_env in
                       range(programDetailsSheet.ncols)]
            for row_index_env in range(2, programDetailsSheet.nrows):
                dictProgramDetails = {
                    keysEnv[col_index_env]: programDetailsSheet.cell(row_index_env, col_index_env).value
                    for col_index_env in range(programDetailsSheet.ncols)}
                programName = dictProgramDetails['Title of the Program']
                isProgramnamePresent = False
                if programName == "":
                    isProgramnamePresent = False
                else:
                    isProgramnamePresent = True
                print(programName)
                scopeEntityType = scopeEntityType
                print(scopeEntityType)
                userEntity = dictProgramDetails['Targeted state at program level'].lstrip().rstrip().split(
                    ",") if \
                    dictProgramDetails['Targeted state at program level'] else terminatingMessage("\"scope_entity\" must not be Empty in \"details\" sheet")
        if sheetEnv.strip().lower() == 'resource details':
            print("--->Checking Resource Details sheet...")
            messageArr = []
            messageArr.append("--->Checking Resource Details sheet...")
            detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
            keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                       range(detailsEnvSheet.ncols)]
            for row_index_env in range(2, detailsEnvSheet.nrows):
                millisecond = int(time.time() * 1000)
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                  for
                                  col_index_env in range(detailsEnvSheet.ncols)}
                resourceNamePGM = dictDetailsEnv['Name of resources in program'] if dictDetailsEnv['Name of resources in program'] else terminatingMessage("\"Name of resources in program\" must not be Empty in \"Resource Details\" sheet")
                resourceTypePGM = dictDetailsEnv['Type of resources'] if dictDetailsEnv['Type of resources'] else terminatingMessage("\"Type of resources\" must not be Empty in \"Resource Details\" sheet")
                resourceLinkOrExtPGM = dictDetailsEnv['Resource Link'] if dictDetailsEnv['Resource Link'] else terminatingMessage("\"Resource Link\" must not be Empty in \"Resource Details\" sheet")
                if str(dictDetailsEnv['Type of resources']).lower().strip() == "course":
                    # isCourse = True
                    isCourse = False
                else:
                    isCourse = False
                    resourceStatus = dictDetailsEnv['Resource Status'] if dictDetailsEnv['Resource Status'] else terminatingMessage("\"Resource Status\" must not be Empty in \"Resource Details\" sheet")
                    if resourceStatus.strip()=="New Upload":
                        print("--->Resource Name : "+str(resourceNamePGM))
                        resourceLinkOrExtPGM = str(resourceLinkOrExtPGM).split('/')[5]
                        file_url = 'https://docs.google.com/spreadsheets/d/' + resourceLinkOrExtPGM + '/export?format=xlsx'
                        if not os.path.isdir('InputFiles'):
                            os.mkdir('InputFiles')
                        dest_file = 'InputFiles'
                        addObservationSolution = wget.download(file_url, dest_file)
                        print("--->solution input file successfully downloaded" + str(addObservationSolution))
                        mainFunc(MainFilePath, programFile, addObservationSolution, millisecond, isProgramnamePresent,isCourse, )
end_time = time.time()
print("Execution time in sec : " + str(end_time - start_time))




